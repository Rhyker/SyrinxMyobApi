import configparser
import os
import sys

# OPEN CONFIGURATION FILE
config = configparser.ConfigParser()
curpath = os.path.dirname(os.path.realpath(sys.argv[0]))
cfgpath = os.path.join(curpath, 'settings.ini')
config.read(cfgpath)


def convert_log_to_excel(log_file_sum, log_file_det, date):
    import openpyxl
    import datetime
    from openpyxl.styles import Font

    # INTEGER REPRESENTATION OF DAY
    day_number = {"Monday": 0,
                  "Tuesday": 1,
                  "Wednesday": 2,
                  "Thursday": 3,
                  "Friday": 4,
                  "Saturday": 5,
                  "Sunday": 6}

    directory = config.get("DEFAULT", "folder_location")
    email_day = day_number[config.get("EMAIL", "send_on_day")]

    # OPEN TEMPLATE
    templatepath = os.path.join(curpath, 'LogTemplate.xlsx')
    template_ex = templatepath

    # WRITE TO TEMPLATE
    wb = openpyxl.load_workbook(template_ex)
    ws = wb.active

    today = datetime.date.today().strftime("%d/%m/%Y")

    # Sets date at top
    ws.cell(column=4, row=1, value=today)

    # LOG FILE SUMMARY
    # Inserts data via iteration and changes font
    for row, row_entries in enumerate(log_file_sum.items(), start=7):
        for column, value in enumerate(row_entries, start=1):
            if type(value) is list:
                for columns, items in enumerate(value, start=2):
                    ws.cell(column=columns, row=row, value=items)
                    ws.cell(column=columns, row=row).font = Font(
                        name='Arial Narrow', size=11)
            else:
                ws.cell(column=column, row=row, value=value)
                ws.cell(column=column, row=row).font = Font(
                    name='Arial Narrow', size=11)

    # LOG FILE DETAILS
    ws = wb["Detailed"]
    ws.cell(column=6, row=1, value=today)

    # Inserts data via iteration and changes font
    for row, row_entries in enumerate(log_file_det.items(), start=7):
        for column, value in enumerate(row_entries, start=1):
            if type(value) is list:
                for columns, items in enumerate(value, start=2):
                    ws.cell(column=columns, row=row, value=items)
                    ws.cell(column=columns, row=row).font = Font(
                        name='Arial Narrow', size=11)

            else:
                ws.cell(column=column, row=row, value=value)
                ws.cell(column=column, row=row).font = Font(
                    name='Arial Narrow', size=11)

    # SAVE THE FILE
    filename = r'Overdue_Log_' + date + '.xlsx'
    directory_filename = directory + '\\' + filename
    wb.save(directory_filename)

    if datetime.datetime.now().weekday() == email_day:
        email_log(attachment_file=directory_filename, filename=filename)


def email_log(attachment_file=None,
              filename="SyrinxLog",
              body=config.get("EMAIL", "body").replace("\\n", "\n"),
              subject=config.get("EMAIL", "subject")):
    import smtplib
    import ssl
    from email.message import EmailMessage
    context = ssl.create_default_context()

    server = config.get("EMAIL", "server")

    message = EmailMessage()
    message.set_content(body)
    message["Subject"] = subject
    message["From"] = config.get("EMAIL", "login_id")

    message["To"] = config.get("EMAIL", "recipients").split(",")

    # TODO: Will need to add functionality to detect attachment type if needed for other purposes
    try:
        with open(attachment_file, 'rb') as f:
            file_data = f.read()
        message.add_attachment(file_data, maintype="application",
                               subtype="xlsx", filename=filename)
    except Exception as e:
        print(e)

    # Send the email
    server = smtplib.SMTP(server, port=587)
    server.starttls(context=context)
    server.login(config.get("EMAIL", "login_id"),
                 config.get("EMAIL", "login_pw"))
    server.send_message(message)
    server.quit()
